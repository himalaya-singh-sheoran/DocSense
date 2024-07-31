import openpyxl
from openpyxl.styles import Font, PatternFill

# Create a new workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active

# Define the colors for "red", "amber", and "green"
color_dict = {
    "red": "FF0000",
    "amber": "FFC000",
    "green": "00FF00"
}

# Input values to be written in the Excel cells
input_values = ["red", "amber", "green", "red", "green"]

# Use the Unicode character for a dot and set the font
dot_char = "●"  # Unicode character for a large dot
font = Font(name='Calibri', size=20)  # Adjust size as needed

# Write input values as dots in cells and apply corresponding color fill
for row, value in enumerate(input_values, start=1):
    cell = ws.cell(row=row, column=1, value=dot_char)
    cell.font = font
    if value in color_dict:
        fill = PatternFill(start_color=color_dict[value], end_color=color_dict[value], fill_type="solid")
        cell.fill = fill

# Save the workbook
wb.save("colored_dots.xlsx")
